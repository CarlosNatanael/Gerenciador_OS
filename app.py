from datetime import datetime
from werkzeug.security import check_password_hash, generate_password_hash 
from functools import wraps
from flask import g
from flask import send_file
from flask import Flask, render_template, request, redirect, url_for, session, flash
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sqlite3
import io


def manutencao_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('tipo') != 'manutencao':
            flash('Acesso não autorizado', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Configuração do banco de dados (mantenha apenas uma versão)
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect('db/os.db')
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA busy_timeout=5000")
    return g.db

def close_db(e=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

app = Flask(__name__)
app.secret_key = 'o8QYV$%7D&(q'

@app.route('/admin/relatorio/os')
def gerar_relatorio_os():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    try:
        conn = get_db()
        cursor = conn.cursor()

        # Query para buscar todas as OS com informações relevantes
        # Incluindo nome do solicitante e uma lista concatenada de técnicos participantes
        # A subquery para técnicos participantes pode ser um pouco complexa em SQLite puro
        # Vamos buscar os dados principais primeiro e depois os participantes por OS se necessário,
        # ou usar uma query mais elaborada se o seu SQLite suportar bem GROUP_CONCAT com JOINs.

        cursor.execute("""
            SELECT 
                os.id, os.equipamento, os.problema, os.prioridade, os.status,
                os.data as data_abertura, os.inicio as data_inicio_reparo, os.fim as data_conclusao,
                os.local, os.setor, os.solucao, os.tempo_reparo,
                u_solicitante.nome as nome_solicitante,
                u_tecnico_sistema.nome as nome_tecnico_sistema -- Usuário que agendou/iniciou no sistema
            FROM ordens_servico os
            LEFT JOIN usuarios u_solicitante ON os.solicitante_id = u_solicitante.id
            LEFT JOIN usuarios u_tecnico_sistema ON os.tecnico_id = u_tecnico_sistema.id
            ORDER BY os.id DESC
        """)
        ordens_servico = cursor.fetchall()

        # Criar um Workbook Excel em memória
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatorio OS"

        # Definir Cabeçalhos
        headers = [
            "ID OS", "Equipamento", "Problema", "Prioridade", "Status", 
            "Data Abertura", "Data Início Reparo", "Data Conclusão",
            "Local", "Setor", "Solicitante", "Técnico Sistema (Agendou/Iniciou)",
            "Técnicos Participantes (Reparo)", "Solução", "Tempo Reparo (min)"
        ]
        sheet.append(headers)

        # Estilizar Cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004085", end_color="004085", fill_type="solid") # Azul escuro
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            # Ajustar largura da coluna (aproximado)
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 20 if len(header_title) < 15 else len(header_title) * 1.2


        # Popular dados das OS
        for row_num, os_data_row in enumerate(ordens_servico, 2): # Começa da linha 2
            os_dict = dict(os_data_row) # Converter para dicionário para fácil acesso

            # Buscar técnicos participantes para esta OS específica
            cursor.execute("""
                SELECT t.nome
                FROM participantes_os po
                JOIN tecnicos t ON po.tecnico_ref_id = t.id
                WHERE po.os_id = ?
                ORDER BY t.nome
            """, (os_dict['id'],))
            participantes_db = cursor.fetchall()
            nomes_participantes = ", ".join([p['nome'] for p in participantes_db]) if participantes_db else "N/A"

            row_data = [
                os_dict.get('id'),
                os_dict.get('equipamento'),
                os_dict.get('problema'),
                os_dict.get('prioridade'),
                os_dict.get('status'),
                os_dict.get('data_abertura'),
                os_dict.get('data_inicio_reparo', ''), # Default para string vazia se None
                os_dict.get('data_conclusao', ''),   # Default para string vazia se None
                os_dict.get('local', ''),
                os_dict.get('setor', ''),
                os_dict.get('nome_solicitante', ''),
                os_dict.get('nome_tecnico_sistema', ''),
                nomes_participantes,
                os_dict.get('solucao', ''),
                os_dict.get('tempo_reparo', '')
            ]
            sheet.append(row_data)
            
            # Aplicar borda às células de dados
            for col_num in range(1, len(headers) + 1):
                sheet.cell(row=row_num, column=col_num).border = thin_border


        # Salvar o workbook em um stream de bytes
        excel_stream = io.BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0) # Voltar ao início do stream

        # conn.close() # Gerenciado pelo teardown_appcontext

        # Nome do arquivo para download
        filename = f"relatorio_os_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"Erro ao gerar relatório OS: {str(e)}") # Log do erro no console do Flask
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))

# Rota inicial
@app.route('/')
def index():
    if 'usuario' in session:
        tipo_usuario = session.get('tipo')
        if tipo_usuario == 'admin' or tipo_usuario == 'master-admin':
            return redirect(url_for('admin_dashboard'))
        elif tipo_usuario == 'manutencao':
            return redirect(url_for('manutencao_dashboard'))
        elif tipo_usuario == 'solicitante':
            return redirect(url_for('minhas_os')) 
        else:
            session.clear()
            flash('Tipo de usuário inválido na sessão.', 'warning')
            return redirect(url_for('login'))
        
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario_form = request.form.get('usuario', '').strip()
        senha_form = request.form.get('senha', '')
        
        print(f"\nDEBUG: Tentativa de login para usuário: [{usuario_form}]") # DEBUG

        if not usuario_form or not senha_form:
            flash('Usuário e senha são obrigatórios.', 'warning')
            print("DEBUG: Login falhou - Usuário ou senha vazios.") # DEBUG
            return render_template('login.html')

        conn = get_db()
        cursor = conn.cursor()
        
        print(f"DEBUG: Consultando DB por usuário: {usuario_form}") # DEBUG
        cursor.execute("SELECT * FROM usuarios WHERE usuario = ? AND ativo = 1", (usuario_form,))
        user_data = cursor.fetchone()

        if user_data:
            print(f"DEBUG: Usuário encontrado no DB: ID={user_data['id']}, Nome={user_data['nome']}, Tipo={user_data['tipo']}, Ativo={user_data['ativo']}") # DEBUG
            is_password_correct = check_password_hash(user_data['senha'], senha_form)
            print(f"DEBUG: Verificação de senha para [{usuario_form}]: {is_password_correct}") # DEBUG

            if is_password_correct:
                session['usuario'] = user_data['usuario']
                session['tipo'] = user_data['tipo']
                session['user_id'] = user_data['id']
                print(f"DEBUG: Sessão configurada: {dict(session)}") # DEBUG

                if user_data['tipo'] == 'solicitante':
                    print("DEBUG: Redirecionando para 'minhas_os' (solicitante)\n") # DEBUG
                    return redirect(url_for('minhas_os'))
                elif user_data['tipo'] == 'manutencao':
                    print("DEBUG: Redirecionando para 'manutencao_dashboard' (manutencao)\n") # DEBUG
                    return redirect(url_for('manutencao_dashboard'))
                elif user_data['tipo'] == 'admin':
                    print("DEBUG: Redirecionando para 'admin_dashboard' (admin)\n") # DEBUG
                    return redirect(url_for('admin_dashboard'))
                elif user_data['tipo'] == 'master-admin':
                    print("DEBUG: Redirecionando para 'admin_dashboard' (master-admin)\n") # DEBUG
                    return redirect(url_for('admin_dashboard')) 
                else:
                    flash('Tipo de usuário desconhecido ou não autorizado.', 'danger')
                    print(f"DEBUG: Login falhou - Tipo de usuário desconhecido: [{user_data['tipo']}]") # DEBUG
                    session.clear()
                    return redirect(url_for('login'))
            else:
                flash('Usuário ou senha inválidos, ou usuário inativo.', 'danger')
                print("DEBUG: Login falhou - Senha incorreta.\n") # DEBUG
        else:
            flash('Usuário ou senha inválidos, ou usuário inativo.', 'danger')
            print(f"DEBUG: Login falhou - Usuário [{usuario_form}] não encontrado ou inativo.\n") # DEBUG
            
    print("\nDEBUG: Renderizando página de login (login.html).\n") # DEBUG
    return render_template('login.html')

# Rota logout
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# --- SOLICITANTE ---
@app.route('/solicitante/abrir', methods=['GET', 'POST'])
def abrir_os():
    if session.get('tipo') not in ['solicitante', 'admin', 'master-admin']:
        flash('Acesso não autorizado para esta funcionalidade.', 'danger')
        return redirect(url_for('login'))

    conn_get = get_db() # Para buscar locais no GET
    cursor_get = conn_get.cursor()
    cursor_get.execute("SELECT id, nome FROM locais WHERE ativo = 1 ORDER BY nome")
    locais_ativos = cursor_get.fetchall()

    if request.method == 'POST':
        equipamento = request.form.get('equipamento', '').strip()
        problema = request.form.get('problema', '').strip()
        prioridade = request.form.get('prioridade') # Será 'normal' ou 'urgente'
        local_selecionado = request.form.get('local')
        setor = request.form.get('setor', '').strip()
        
        
        if not equipamento or not problema or not local_selecionado or not prioridade:
            flash('Preencha todos os campos obrigatórios (Equipamento, Local, Problema e Prioridade).', 'danger')
            return render_template('solicitante/abrir_os.html', 
                                   locais=locais_ativos, 
                                   datetime=datetime,
                                   request_form_data=request.form)
        
        data_abertura = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        try:
            cursor_get.execute("SELECT id FROM usuarios WHERE usuario = ?", (session['usuario'],))
            user = cursor_get.fetchone()

            if not user:
                flash("Erro de sessão do usuário. Faça login novamente.", "danger")
                return redirect(url_for('login'))

            cursor_get.execute("""
                INSERT INTO ordens_servico 
                (data, equipamento, problema, prioridade, status, solicitante_id, local, setor) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (data_abertura, equipamento, problema, prioridade, 'Aberta', user['id'], local_selecionado, setor))
            
            conn_get.commit()
            flash('Ordem de serviço enviada com sucesso! A equipe de manutenção fará o agendamento.', 'success') # Mensagem ajustada
            
            if session.get('tipo') == 'solicitante':
                return redirect(url_for('minhas_os')) 
            else: # admin ou master-admin
                return redirect(url_for('admin_dashboard'))
            
        except sqlite3.Error as e:
            conn_get.rollback()
            flash(f'Erro ao salvar OS: {str(e)}', 'danger')

        return render_template('solicitante/abrir_os.html', 
                               locais=locais_ativos, 
                               datetime=datetime, # Se usado no template
                               request_form_data=request.form) # Para repopular

    return render_template('solicitante/abrir_os.html', 
                           locais=locais_ativos, 
                           datetime=datetime)

@app.route('/solicitante/minhas_os')
def minhas_os():
    if session.get('tipo') != 'solicitante':
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    
    # Busca o ID do solicitante
    cursor.execute("SELECT id FROM usuarios WHERE usuario = ?", (session['usuario'],))
    user = cursor.fetchone()
    
    # Busca todas as OSs do solicitante que não estão concluídas
    cursor.execute("""
        SELECT os.*, u.nome as tecnico_nome
        FROM ordens_servico os
        LEFT JOIN usuarios u ON os.tecnico_id = u.id
        WHERE os.solicitante_id = ? AND os.status != 'Concluída'
        ORDER BY os.data DESC
    """, (user['id'],))
    os_abertas = cursor.fetchall()
    
    conn.close()

    return render_template('solicitante/minhas_os.html', os_abertas=os_abertas)

# --- MANUTENÇÃO ---
@app.route('/manutencao')
@manutencao_required
def manutencao_dashboard():
    if session.get('tipo') != 'manutencao':
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ordens_servico WHERE status != 'Concluída'")
    os_abertas = cursor.fetchall()
    conn.close()

    return render_template('manutencao/listar_os.html', os_abertas=os_abertas)

@app.route('/manutencao/concluir/<int:id_os>', methods=['POST'])
@manutencao_required
def concluir_os(id_os):
    if request.method == 'POST':
        solucao = request.form.get('solucao','').strip()
        data_conclusao_str = request.form.get('data_conclusao_manual')
        hora_conclusao_str = request.form.get('hora_conclusao_manual')
        # ids_tecnicos_participantes são os IDs da tabela 'tecnicos'
        ids_tecnicos_participantes = request.form.getlist('tecnicos_participantes') 

        if not solucao or not data_conclusao_str or not hora_conclusao_str or not ids_tecnicos_participantes:
            flash('Solução, data/hora da conclusão e pelo menos um técnico participante são obrigatórios.', 'danger')
            # É importante passar 'todos_tecnicos_manutencao' de volta se o template precisar para repopular
            # Vamos buscar novamente para garantir
            conn_err = get_db()
            cursor_err = conn_err.cursor()
            cursor_err.execute("SELECT id, nome, tipo_tecnico as especialidade FROM tecnicos WHERE ativo = 1 ORDER BY nome")
            todos_tecnicos_err = cursor_err.fetchall()
            
            cursor_err.execute("SELECT * FROM ordens_servico WHERE id = ?", (id_os,))
            os_data_err = cursor_err.fetchone()

            return render_template('manutencao/detalhe_os.html', 
                                   os=dict(os_data_err) if os_data_err else None, 
                                   todos_tecnicos_manutencao=todos_tecnicos_err,
                                   # Você também precisa passar 'solicitante', 'tecnico' (sistema), 'datetime' se o template os usa
                                   # Adicione-os aqui se necessário, buscando-os novamente.
                                   request_form_data=request.form, # Para repopular
                                   datetime=datetime # Para os valores padrão de data/hora
                                   )


        try:
            fim_manual_dt_str = f"{data_conclusao_str} {hora_conclusao_str}:00"
            fim_manual_dt = datetime.strptime(fim_manual_dt_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            flash('Formato de data ou hora da conclusão inválido.', 'danger')
            # Mesma lógica de repopulação que acima
            conn_err = get_db()
            cursor_err = conn_err.cursor()
            cursor_err.execute("SELECT id, nome, tipo_tecnico as especialidade FROM tecnicos WHERE ativo = 1 ORDER BY nome")
            todos_tecnicos_err = cursor_err.fetchall()
            cursor_err.execute("SELECT * FROM ordens_servico WHERE id = ?", (id_os,))
            os_data_err = cursor_err.fetchone()
            return render_template('manutencao/detalhe_os.html', 
                                   os=dict(os_data_err) if os_data_err else None, 
                                   todos_tecnicos_manutencao=todos_tecnicos_err,
                                   request_form_data=request.form,
                                   datetime=datetime)


        conn = get_db()
        cursor = conn.cursor()

        try:
            cursor.execute("SELECT data, inicio FROM ordens_servico WHERE id = ?", (id_os,))
            os_data_db = cursor.fetchone()

            if not os_data_db:
                flash('OS não encontrada.', 'danger')
                return redirect(url_for('manutencao_dashboard'))

            inicio_reparo_str = os_data_db['inicio'] if os_data_db['inicio'] else os_data_db['data']
            inicio_reparo_dt = datetime.strptime(inicio_reparo_str, '%Y-%m-%d %H:%M:%S')

            if fim_manual_dt < inicio_reparo_dt:
                flash('A data de conclusão não pode ser anterior à data de início/abertura da OS.', 'warning')
                # Mesma lógica de repopulação que acima
                cursor.execute("SELECT id, nome, tipo_tecnico as especialidade FROM tecnicos WHERE ativo = 1 ORDER BY nome")
                todos_tecnicos_err = cursor.fetchall()
                cursor.execute("SELECT * FROM ordens_servico WHERE id = ?", (id_os,))
                os_data_err = cursor.fetchone()
                return render_template('manutencao/detalhe_os.html', 
                                   os=dict(os_data_err) if os_data_err else None, 
                                   todos_tecnicos_manutencao=todos_tecnicos_err,
                                   request_form_data=request.form,
                                   datetime=datetime)


            tempo_total_segundos = (fim_manual_dt - inicio_reparo_dt).total_seconds()
            tempo_total_minutos = round(tempo_total_segundos / 60)
            inicio_final_para_db = inicio_reparo_str 
            
            cursor.execute("""
                UPDATE ordens_servico
                SET solucao = ?, 
                    status = 'Concluída', 
                    tempo_reparo = ?, 
                    fim = ?,
                    inicio = ?
                WHERE id = ?
            """, (solucao, tempo_total_minutos, fim_manual_dt.strftime('%Y-%m-%d %H:%M:%S'), inicio_final_para_db, id_os))
            
            cursor.execute("DELETE FROM participantes_os WHERE os_id = ?", (id_os,))
            
            for tecnico_id_str in ids_tecnicos_participantes:
                try:
                    tecnico_individual_id = int(tecnico_id_str) # Este é o ID da tabela 'tecnicos'
                    cursor.execute("""
                        INSERT INTO participantes_os (os_id, tecnico_ref_id) -- CORRIGIDO AQUI
                        VALUES (?, ?)
                    """, (id_os, tecnico_individual_id))
                except ValueError:
                    flash(f"ID de técnico inválido encontrado: {tecnico_id_str}", "warning")
            
            cursor.execute("""
                INSERT INTO historico_os (os_id, usuario_id, acao, observacao)
                VALUES (?, ?, ?, ?)
            """, (id_os, session.get('user_id'), 'OS Concluída', f"Solução: {solucao}"))
            
            conn.commit()
            flash('OS concluída com sucesso!', 'success')
            return redirect(url_for('detalhe_os', id_os=id_os))

        except sqlite3.Error as e:
            conn.rollback()
            flash(f'Erro de banco de dados ao concluir OS: {str(e)}', 'danger')
        except Exception as e:
            flash(f'Erro ao concluir OS: {str(e)}', 'danger')
        
        # Se chegou aqui (erro no try), precisa repopular o formulário de detalhe_os
        # Buscando todos os dados novamente
        cursor.execute("SELECT id, nome, tipo_tecnico as especialidade FROM tecnicos WHERE ativo = 1 ORDER BY nome")
        todos_tecnicos_repop = cursor.fetchall()
        cursor.execute("SELECT * FROM ordens_servico WHERE id = ?", (id_os,))
        os_data_repop = cursor.fetchone()
        # Você também precisaria buscar solicitante_info, tecnico_sistema_info, historico para o detalhe_os.html
        # Esta parte de repopulação pode ficar complexa. O ideal é que as validações iniciais capturem a maioria dos erros.
        return render_template('manutencao/detalhe_os.html', 
                               os=dict(os_data_repop) if os_data_repop else None, 
                               todos_tecnicos_manutencao=todos_tecnicos_repop,
                               request_form_data=request.form,
                               datetime=datetime) # Adicione outras variáveis que detalhe_os.html espera

    return redirect(url_for('manutencao_dashboard'))

@app.route('/manutencao/iniciar/<int:id>')
def iniciar_os(id):
    if session.get('tipo') != 'manutencao':
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Verifica se está agendada
        cursor.execute("SELECT status FROM ordens_servico WHERE id = ?", (id,))
        os_status = cursor.fetchone()['status']
        
        if os_status != 'Agendada':
            flash('Só é possível iniciar OSs agendadas', 'warning')
            return redirect(url_for('manutencao_dashboard'))
        
        # Atualiza status
        cursor.execute("""
            UPDATE ordens_servico 
            SET status = 'Em andamento',
                inicio = datetime('now')
            WHERE id = ?
        """, (id,))
        
        # Registra no histórico
        cursor.execute("""
            INSERT INTO historico_os 
            (os_id, usuario_id, acao)
            VALUES (?, ?, ?)
        """, (id, session.get('user_id'), 'Reparo iniciado'))
        
        conn.commit()
        flash('Reparo iniciado com sucesso!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao iniciar reparo: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('manutencao_dashboard'))

@app.route('/manutencao/os/<int:id_os>')
@manutencao_required
def detalhe_os(id_os):
    conn = get_db()
    cursor = conn.cursor()

    try:
        # Busca a OS principal e informações do solicitante e do usuário que agiu como técnico no sistema
        cursor.execute("""
            SELECT os.*, 
                   s.nome as solicitante_nome,
                   s.email as solicitante_email,
                   u_tecnico_sistema.nome as nome_tecnico_sistema  -- Usuário do sistema (ex: 'manutencao')
            FROM ordens_servico os
            LEFT JOIN usuarios s ON os.solicitante_id = s.id
            LEFT JOIN usuarios u_tecnico_sistema ON os.tecnico_id = u_tecnico_sistema.id 
            WHERE os.id = ?
        """, (id_os,))
        os_data = cursor.fetchone()

        if not os_data:
            flash('Ordem de Serviço não encontrada.', 'danger')
            return redirect(url_for('manutencao_dashboard'))

        # Converte a linha da OS para um dicionário para fácil acesso no template
        os_dict = dict(os_data) 
        
        solicitante_info = {
            'nome': os_data['solicitante_nome'],
            'email': os_data['solicitante_email']
        } if os_data['solicitante_nome'] else None

        # O 'tecnico' aqui é o usuário do sistema que agendou/iniciou (ex: o login genérico 'manutencao')
        # Ele não tem 'especialidade' na tabela 'usuarios'.
        tecnico_sistema_info = {
            'nome': os_data['nome_tecnico_sistema']
        } if os_data['nome_tecnico_sistema'] else None


        # Busca histórico de alterações da OS
        cursor.execute("""
            SELECT h.*, u.nome as usuario_nome
            FROM historico_os h
            JOIN usuarios u ON h.usuario_id = u.id
            WHERE h.os_id = ?
            ORDER BY h.data_alteracao DESC
        """, (id_os,))
        historico = cursor.fetchall()

        # Busca todos os TÉCNICOS INDIVIDUAIS ativos da tabela 'tecnicos' para o formulário de conclusão
        todos_tecnicos_individuais = []
        if os_dict['status'] != 'Concluída' and os_dict['status'] != 'Cancelada':
            cursor.execute("""
                SELECT id, nome, tipo_tecnico 
                FROM tecnicos 
                WHERE ativo = 1 
                ORDER BY nome
            """)
            # Renomear 'tipo_tecnico' para 'especialidade' no dicionário para manter consistência com o template antigo, se desejado
            # Ou alterar o template para usar 'tipo_tecnico'
            todos_tecnicos_individuais = [
                {'id': row['id'], 'nome': row['nome'], 'especialidade': row['tipo_tecnico']} 
                for row in cursor.fetchall()
            ]


        # Se a OS já está concluída, busca os TÉCNICOS INDIVIDUAIS que participaram
        participantes_conclusao = []
        if os_dict['status'] == 'Concluída':
            cursor.execute("""
                SELECT t.id, t.nome, t.tipo_tecnico 
                FROM participantes_os po
                JOIN tecnicos t ON po.tecnico_ref_id = t.id
                WHERE po.os_id = ?
                ORDER BY t.nome
            """, (id_os,))
            # Renomear 'tipo_tecnico' para 'especialidade' no dicionário
            participantes_conclusao = [
                {'id': row['id'], 'nome': row['nome'], 'especialidade': row['tipo_tecnico']}
                for row in cursor.fetchall()
            ]
        
        return render_template(
            'manutencao/detalhe_os.html',
            os=os_dict,
            solicitante=solicitante_info,
            tecnico=tecnico_sistema_info, # Passando o usuário do sistema que agiu
            historico=historico,
            todos_tecnicos_manutencao=todos_tecnicos_individuais, # Nome da variável como o template espera
            participantes_conclusao=participantes_conclusao,
            datetime=datetime # Se o template ainda usa para valores padrão de data/hora
        )

    except Exception as e:
        print(f"Erro ao carregar detalhes da OS {id_os}: {str(e)}") 
        flash(f'Erro ao carregar detalhes da OS: {str(e)}', 'danger')
        return redirect(url_for('manutencao_dashboard'))



@app.route('/manutencao/agendar/<int:id_os>', methods=['GET', 'POST'])
@manutencao_required
def agendar_os(id_os):
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM ordens_servico WHERE id = ?", (id_os,))
    os_data = cursor.fetchone()

    if not os_data:
        flash('Ordem de Serviço não encontrada.', 'danger')
        return redirect(url_for('manutencao_dashboard'))

    if os_data['status'] in ['Concluída', 'Cancelada']:
        flash(f"OS #{id_os} já está {os_data['status'].lower()} e não pode ser reagendada.", 'warning')
        return redirect(url_for('detalhe_os', id_os=id_os))

    if request.method == 'POST':
        data_agendamento_str = request.form.get('data_agendamento')
        horario_agendamento_str = request.form.get('horario_agendamento')
        
        # REMOVIDO: Leitura de ids_tecnicos_participantes
        # ids_tecnicos_participantes = request.form.getlist('tecnicos_participantes')

        # Validação apenas para data e horário
        if not data_agendamento_str or not horario_agendamento_str:
            flash('Data e horário são obrigatórios para o agendamento.', 'danger')
            # Não é mais necessário passar 'todos_tecnicos_manutencao' aqui
            return render_template('manutencao/agendar_os.html', 
                                   os=os_data, 
                                   request_form_data=request.form) # Para repopular data/hora
        try:
            # Atualiza a OS com agendamento.
            # O campo 'tecnico_id' na tabela 'ordens_servico' ainda será preenchido
            # com o ID do usuário de manutenção que está fazendo o agendamento.
            # Isso pode ser útil para saber quem agendou.
            # Se você não quiser isso, remova a linha 'tecnico_id = ?' e o parâmetro correspondente.
            cursor.execute("""
                UPDATE ordens_servico 
                SET data_agendamento = ?, 
                    horario_agendamento = ?,
                    status = 'Agendada',
                    tecnico_id = ?  
                WHERE id = ?
            """, (data_agendamento_str, horario_agendamento_str, session.get('user_id'), id_os))
            
            # REMOVIDO: Lógica para gerenciar 'participantes_os'
            # A seleção de técnicos participantes foi movida para a etapa de conclusão da OS.

            # Registra no histórico
            observacao_historico = f"Agendado para {data_agendamento_str} {horario_agendamento_str}."
            # REMOVIDO: Lógica para adicionar nomes dos participantes ao histórico aqui

            cursor.execute("""
                INSERT INTO historico_os 
                (os_id, usuario_id, acao, observacao)
                VALUES (?, ?, ?, ?)
            """, (id_os, session.get('user_id'), 'OS Agendada', observacao_historico))
            
            conn.commit()
            flash('OS agendada com sucesso!', 'success')
            return redirect(url_for('detalhe_os', id_os=id_os))

        except sqlite3.Error as e:
            conn.rollback()
            flash(f'Erro de banco de dados ao agendar OS: {str(e)}', 'danger')
        except Exception as e:
            flash(f'Erro ao agendar OS: {str(e)}', 'danger')
        
        # Se chegou aqui, houve erro, re-renderiza
        return render_template('manutencao/agendar_os.html', 
                               os=os_data, 
                               request_form_data=request.form)

    # Método GET
    # Não é mais necessário passar 'todos_tecnicos_manutencao'
    return render_template('manutencao/agendar_os.html', 
                           os=os_data)

@app.route('/manutencao/registros/novo', methods=['GET', 'POST'])
@manutencao_required
def novo_registro_direto():
    conn = get_db()
    cursor = conn.cursor()

    # Buscar todos os TÉCNICOS INDIVIDUAIS ativos da tabela 'tecnicos' para o formulário
    cursor.execute("""
        SELECT id, nome, tipo_tecnico 
        FROM tecnicos 
        WHERE ativo = 1 
        ORDER BY nome
    """)
    # Mapear 'tipo_tecnico' para 'especialidade' para manter consistência com o template,
    # que já espera 'especialidade' (como em detalhe_os.html).
    todos_tecnicos_manutencao = [
        {'id': row['id'], 'nome': row['nome'], 'especialidade': row['tipo_tecnico']}
        for row in cursor.fetchall()
    ]
    # Não feche a conexão aqui se for usá-la no POST ou se houver erro e re-renderizar

    if request.method == 'POST':
        data_execucao_str = request.form.get('data_execucao')
        hora_execucao_str = request.form.get('hora_execucao')
        duracao_minutos_str = request.form.get('duracao_minutos')
        equipamento_afetado = request.form.get('equipamento_afetado','').strip()
        descricao_servico = request.form.get('descricao_servico','').strip()
        observacoes = request.form.get('observacoes','').strip()
        ids_tecnicos_participantes = request.form.getlist('tecnicos_participantes') # Vem da tabela 'tecnicos'
        
        # criado_por_id será o ID do usuário genérico 'manutencao' logado no sistema
        criado_por_id = session.get('user_id') 

        erros = []
        if not data_execucao_str: erros.append("Data da execução é obrigatória.")
        if not hora_execucao_str: erros.append("Hora da execução é obrigatória.")
        if not duracao_minutos_str: erros.append("Duração da operação é obrigatória.")
        if not descricao_servico: erros.append("Descrição do serviço é obrigatória.")
        if not ids_tecnicos_participantes: erros.append("Pelo menos um técnico participante deve ser selecionado.")

        data_execucao_completa_str = None
        if data_execucao_str and hora_execucao_str:
            try:
                data_execucao_completa_str = f"{data_execucao_str} {hora_execucao_str}:00"
                datetime.strptime(data_execucao_completa_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                erros.append("Formato de data ou hora da execução inválido.")
        
        duracao_minutos = None
        if duracao_minutos_str:
            try:
                duracao_minutos = int(duracao_minutos_str)
                if duracao_minutos <= 0:
                    erros.append("Duração da operação deve ser um número positivo.")
            except ValueError:
                erros.append("Duração da operação deve ser um número.")

        if erros:
            for erro in erros:
                flash(erro, 'danger')
            return render_template(
                'manutencao/novo_registro_direto.html',
                todos_tecnicos_manutencao=todos_tecnicos_manutencao, # Já buscados e formatados
                datetime=datetime,
                request_form_data=request.form # Para repopular os campos
            )

        # A conexão 'conn' e 'cursor' já estão abertas e podem ser usadas para o POST
        try:
            cursor.execute("""
                INSERT INTO registros_manutencao_direta 
                (data_execucao, duracao_minutos, equipamento_afetado, descricao_servico, observacoes, criado_por_id, status)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (data_execucao_completa_str, duracao_minutos, equipamento_afetado, descricao_servico, observacoes, criado_por_id, 'Pendente Aprovacao'))
            
            novo_registro_id = cursor.lastrowid

            # Adicionar participantes na tabela 'participantes_registro_direto'
            # que agora referencia 'tecnicos.id' através de 'tecnico_ref_id'
            for tecnico_id_str in ids_tecnicos_participantes:
                try:
                    tecnico_id = int(tecnico_id_str) # Este é o ID da tabela 'tecnicos'
                    cursor.execute("""
                        INSERT INTO participantes_registro_direto (registro_id, tecnico_ref_id)
                        VALUES (?, ?)
                    """, (novo_registro_id, tecnico_id))
                except ValueError:
                    print(f"Aviso: ID de técnico inválido '{tecnico_id_str}' ignorado para registro direto {novo_registro_id}")

            conn.commit()
            flash('Registro de manutenção direta criado com sucesso e pendente de aprovação!', 'success')
            return redirect(url_for('manutencao_dashboard')) 

        except sqlite3.Error as e:
            conn.rollback()
            flash(f'Erro de banco de dados ao salvar o registro: {str(e)}', 'danger')
        except Exception as e:
            flash(f'Erro ao salvar o registro: {str(e)}', 'danger')
        
        # Se chegou aqui (erro no try), renderiza o form novamente
        return render_template(
            'manutencao/novo_registro_direto.html',
            todos_tecnicos_manutencao=todos_tecnicos_manutencao, # Já buscados e formatados
            datetime=datetime,
            request_form_data=request.form # Para repopular
        )

    # Método GET
    return render_template(
        'manutencao/novo_registro_direto.html',
        todos_tecnicos_manutencao=todos_tecnicos_manutencao, # Já buscados e formatados
        datetime=datetime
    )

# --- ADMIN ---
@app.route('/admin') # Ou a URL que você usa para o admin_dashboard
def admin_dashboard():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado ao dashboard administrativo.', 'danger')
        return redirect(url_for('login'))
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Estatísticas de Ordens de Serviço (OS)
        cursor.execute("""
            SELECT 
                COUNT(*) AS total,
                SUM(CASE WHEN status = 'Concluída' THEN 1 ELSE 0 END) AS concluidas,
                SUM(CASE WHEN status = 'Aberta' THEN 1 ELSE 0 END) AS abertas,
                SUM(CASE WHEN status = 'Em andamento' THEN 1 ELSE 0 END) AS em_andamento,
                SUM(CASE WHEN status = 'Agendada' THEN 1 ELSE 0 END) AS agendadas,
                AVG(CASE WHEN status = 'Concluída' AND tempo_reparo IS NOT NULL THEN tempo_reparo ELSE NULL END) AS media_tempo
            FROM ordens_servico
        """)
        stats_os_data = cursor.fetchone()

        stats_os_dict = dict(stats_os_data) if stats_os_data else {
            'total': 0, 'concluidas': 0, 'abertas': 0, 
            'em_andamento': 0, 'agendadas': 0, 'media_tempo': 0.0
        }
        if stats_os_dict['media_tempo'] is None: # Tratar None para média
            stats_os_dict['media_tempo'] = 0.0

        # Estatísticas de Registros de Manutenção Direta
        cursor.execute("""
            SELECT 
                COUNT(*) AS total_registros,
                SUM(CASE WHEN status = 'Pendente Aprovacao' THEN 1 ELSE 0 END) AS pendente_aprovacao,
                SUM(CASE WHEN status = 'Concluido' THEN 1 ELSE 0 END) AS concluidos_registros,
                SUM(CASE WHEN status = 'Cancelado' THEN 1 ELSE 0 END) AS cancelados_registros
            FROM registros_manutencao_direta
        """)
        stats_registros_data = cursor.fetchone()
        
        stats_registros_dict = dict(stats_registros_data) if stats_registros_data else {
            'total_registros': 0, 'pendente_aprovacao': 0, 
            'concluidos_registros': 0, 'cancelados_registros': 0
        }
        # Garante que os campos existam mesmo que a query não retorne nada ou algum SUM seja NULL
        stats_registros_dict.setdefault('total_registros', 0)
        stats_registros_dict.setdefault('pendente_aprovacao', 0)
        stats_registros_dict.setdefault('concluidos_registros', 0)
        stats_registros_dict.setdefault('cancelados_registros', 0)


        # Últimas Ordens de Serviço para exibição
        cursor.execute("""
            SELECT os.*, u.nome as solicitante_nome
            FROM ordens_servico os
            JOIN usuarios u ON os.solicitante_id = u.id
            ORDER BY os.data DESC
            LIMIT 10 
        """)
        todas_os = cursor.fetchall()
        
        # Combina as estatísticas
        stats_completas = {
            'os': stats_os_dict,
            'registros_diretos': stats_registros_dict
        }

        return render_template(
            'administrador/dashboard.html', 
            stats=stats_completas, # Passa o dicionário combinado
            todas_os=todas_os
        )

    except sqlite3.Error as e:
        print(f"Erro de banco de dados no admin_dashboard: {str(e)}")
        flash(f'Erro ao carregar dados do dashboard: {str(e)}', 'danger')
        # Em caso de erro, pode ser útil ter um dicionário de stats vazio para evitar erros no template
        stats_vazias = {
            'os': {'total': 0, 'concluidas': 0, 'abertas': 0, 'em_andamento': 0, 'agendadas': 0, 'media_tempo': 0.0},
            'registros_diretos': {'total_registros': 0, 'pendente_aprovacao': 0, 'concluidos_registros': 0, 'cancelados_registros': 0}
        }
        return render_template('administrador/dashboard.html', stats=stats_vazias, todas_os=[])
    except Exception as e:
        print(f"Erro geral no admin_dashboard: {str(e)}")
        flash(f'Ocorreu um erro inesperado: {str(e)}', 'danger')
        stats_vazias = { # Estrutura de fallback
            'os': {'total': 0, 'concluidas': 0, 'abertas': 0, 'em_andamento': 0, 'agendadas': 0, 'media_tempo': 0.0},
            'registros_diretos': {'total_registros': 0, 'pendente_aprovacao': 0, 'concluidos_registros': 0, 'cancelados_registros': 0}
        }
        return render_template('administrador/dashboard.html', stats=stats_vazias, todas_os=[])

@app.route('/admin/os/<int:id_os>')
def detalhe_os_admin(id_os):
    if session.get('tipo') not in ['admin', 'master-admin']: # Permissão
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Busca a OS principal com informações do solicitante e do usuário do sistema que agiu como técnico
        cursor.execute("""
            SELECT os.*, 
                   s.nome as solicitante_nome,
                   s.email as solicitante_email,
                   u_tecnico_sistema.nome as nome_tecnico_sistema 
            FROM ordens_servico os
            LEFT JOIN usuarios s ON os.solicitante_id = s.id
            LEFT JOIN usuarios u_tecnico_sistema ON os.tecnico_id = u_tecnico_sistema.id
            WHERE os.id = ?
        """, (id_os,))
        os_data = cursor.fetchone()
        
        if not os_data:
            flash('Ordem de Serviço não encontrada.', 'danger')
            return redirect(url_for('admin_dashboard'))
        
        # Converte a linha da OS para um dicionário
        os_dict = dict(os_data)
        
        solicitante_info = {
            'nome': os_data['solicitante_nome'],
            'email': os_data['solicitante_email']
        } if os_data['solicitante_nome'] else None
        
        tecnico_sistema_info = { # Usuário do sistema que agendou/iniciou a OS (ex: login 'manutencao')
            'nome': os_data['nome_tecnico_sistema']
        } if os_data['nome_tecnico_sistema'] else None
        
        # Busca os TÉCNICOS INDIVIDUAIS participantes desta OS da tabela 'tecnicos'
        # através da tabela de ligação 'participantes_os'
        participantes_individuais = []
        cursor.execute("""
            SELECT t.id, t.nome, t.tipo_tecnico  -- Seleciona tipo_tecnico da tabela tecnicos
            FROM participantes_os po
            JOIN tecnicos t ON po.tecnico_ref_id = t.id  -- Junta com 'tecnicos' usando 'tecnico_ref_id'
            WHERE po.os_id = ?
            ORDER BY t.nome
        """, (id_os,))
        
        # Mapeia 'tipo_tecnico' para 'funcao' para o template do admin, que espera 'part.funcao'
        participantes_individuais = [
            {'id': row['id'], 
             'nome': row['nome'], 
             'email': None, # Tabela 'tecnicos' não tem email. Se necessário, ajuste ou remova do template.
             'funcao': row['tipo_tecnico'].capitalize() if row['tipo_tecnico'] else 'Técnico'} 
            for row in cursor.fetchall()
        ]

        # Busca histórico da OS
        historico = []
        cursor.execute("""
            SELECT h.*, u.nome as usuario_nome
            FROM historico_os h
            JOIN usuarios u ON h.usuario_id = u.id
            WHERE h.os_id = ?
            ORDER BY h.data_alteracao DESC
        """, (id_os,))
        historico = cursor.fetchall()
        
        return render_template(
            'administrador/detalhes_os.html', 
            os=os_dict,
            solicitante=solicitante_info,
            tecnico=tecnico_sistema_info, 
            participantes=participantes_individuais, 
            historico=historico
        )
        
    except sqlite3.Error as e:
        print(f"Erro de banco de dados ao acessar OS {id_os} (admin): {str(e)}")
        flash(f'Erro de banco de dados ao carregar detalhes da OS: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        print(f"Erro geral ao acessar OS {id_os} (admin): {str(e)}")
        flash(f'Erro ao carregar detalhes da OS: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))



# --- Registro de Manutenção ----
@app.route('/admin/registros_manutencao')
def listar_registros_diretos():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    
    status_filtro = request.args.get('status_filtro') # Pega o filtro da URL
    
    query = """
        SELECT rmd.*, u_criador.nome as nome_criador
        FROM registros_manutencao_direta rmd
        JOIN usuarios u_criador ON rmd.criado_por_id = u_criador.id
    """
    params = []
    if status_filtro:
        query += " WHERE rmd.status = ?"
        params.append(status_filtro)
    
    query += " ORDER BY rmd.data_registro DESC"
    
    cursor.execute(query, params)
    registros = cursor.fetchall()

    return render_template('administrador/listar_registros_diretos.html', 
                           registros=registros, 
                           status_filtrado=status_filtro)

@app.route('/admin/registros_manutencao/processar/<int:id_registro>', methods=['POST'])
def processar_registro_direto(id_registro):
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    acao = request.form.get('acao') # 'concluir' ou 'cancelar'
    admin_id = session.get('user_id')
    data_atual_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cursor = conn.cursor()

    try:
        if acao == 'concluir':
            cursor.execute("""
                UPDATE registros_manutencao_direta
                SET status = 'Concluido',
                    concluido_por_admin_id = ?,
                    data_conclusao_admin = ?
                WHERE id = ? AND status = 'Pendente Aprovacao' 
            """, (admin_id, data_atual_str, id_registro))
            if cursor.rowcount > 0:
                flash('Registro de manutenção concluído com sucesso!', 'success')
            else:
                flash('Não foi possível concluir o registro (pode já ter sido processado ou não existe).', 'warning')
        
        elif acao == 'cancelar':
            cursor.execute("""
                UPDATE registros_manutencao_direta
                SET status = 'Cancelado',
                    concluido_por_admin_id = ?, 
                    data_conclusao_admin = ? 
                WHERE id = ? AND status = 'Pendente Aprovacao'
            """, (admin_id, data_atual_str, id_registro))
            if cursor.rowcount > 0:
                flash('Registro de manutenção cancelado.', 'info')
            else:
                flash('Não foi possível cancelar o registro (pode já ter sido processado ou não existe).', 'warning')
        else:
            flash('Ação inválida.', 'danger')

        conn.commit()
    except sqlite3.Error as e:
        conn.rollback()
        flash(f'Erro de banco de dados ao processar o registro: {str(e)}', 'danger')
    except Exception as e:
        flash(f'Erro ao processar o registro: {str(e)}', 'danger')
    
    return redirect(url_for('detalhe_registro_direto', id_registro=id_registro))

@app.route('/admin/registros_manutencao/<int:id_registro>')
def detalhe_registro_direto(id_registro):
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()

    try:
        # Buscar o registro principal e o nome de quem criou e quem concluiu (se houver)
        cursor.execute("""
            SELECT rmd.*, 
                   u_criador.nome as nome_criador, 
                   u_admin.nome as nome_admin_conclusao
            FROM registros_manutencao_direta rmd
            JOIN usuarios u_criador ON rmd.criado_por_id = u_criador.id
            LEFT JOIN usuarios u_admin ON rmd.concluido_por_admin_id = u_admin.id
            WHERE rmd.id = ?
        """, (id_registro,))
        registro_data = cursor.fetchone()

        if not registro_data:
            flash('Registro de manutenção não encontrado.', 'danger')
            return redirect(url_for('listar_registros_diretos'))
        
        registro_dict = dict(registro_data) # Converter para dicionário

        # Buscar os TÉCNICOS INDIVIDUAIS participantes deste registro
        # da tabela 'tecnicos' através da tabela de ligação 'participantes_registro_direto'
        participantes_individuais = []
        cursor.execute("""
            SELECT t.id, t.nome, t.tipo_tecnico 
            FROM participantes_registro_direto prd
            JOIN tecnicos t ON prd.tecnico_ref_id = t.id  -- CORRIGIDO: junta com 'tecnicos' usando 'tecnico_ref_id'
            WHERE prd.registro_id = ?
            ORDER BY t.nome
        """, (id_registro,))
        
        # Mapear 'tipo_tecnico' para 'especialidade' para o template, se ele espera esse nome
        # O template 'detalhe_registro_direto.html' usa part.especialidade
        participantes_individuais = [
            {'id': row['id'], 
             'nome': row['nome'], 
             'especialidade': row['tipo_tecnico'].capitalize() if row['tipo_tecnico'] else 'Técnico'} 
            for row in cursor.fetchall()
        ]
        
        return render_template(
            'administrador/detalhe_registro_direto.html', 
            registro=registro_dict, 
            participantes=participantes_individuais
        )

    except sqlite3.Error as e:
        print(f"Erro de banco de dados ao acessar registro direto {id_registro}: {str(e)}")
        flash(f'Erro de banco de dados ao carregar detalhes do registro: {str(e)}', 'danger')
        return redirect(url_for('listar_registros_diretos'))
    except Exception as e:
        print(f"Erro geral ao acessar registro direto {id_registro}: {str(e)}")
        flash(f'Erro ao carregar detalhes do registro: {str(e)}', 'danger')
        return redirect(url_for('listar_registros_diretos'))


# --- ADMIN - Configurações ---
@app.route('/admin/configuracoes')
def admin_configuracoes():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))
    return render_template('administrador/configuracoes.html')

# Rota para listar os técnicos individuais
@app.route('/admin/tecnicos')
def listar_tecnicos():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM tecnicos ORDER BY nome")
    lista_de_tecnicos = cursor.fetchall()
    # conn.close() # Gerenciado pelo teardown_appcontext

    return render_template('administrador/listar_tecnicos.html', tecnicos=lista_de_tecnicos)

# Rota para adicionar novo técnico individual (você já tem esta)
@app.route('/admin/tecnicos/novo', methods=['GET', 'POST'])
def adicionar_tecnico():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo_tecnico = request.form.get('tipo_tecnico')

        if not nome or not tipo_tecnico:
            flash('Nome e Tipo do Técnico são obrigatórios.', 'warning')
            return render_template('administrador/form_tecnico.html', 
                                   tecnico=request.form, 
                                   titulo="Adicionar Novo Técnico",
                                   acao="Adicionar")
        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO tecnicos (nome, tipo_tecnico, ativo) VALUES (?, ?, 1)",
                (nome, tipo_tecnico)
            )
            conn.commit()
            flash('Técnico adicionado com sucesso!', 'success')
            return redirect(url_for('listar_tecnicos'))
        except sqlite3.IntegrityError:
            flash('Já existe um técnico com este nome.', 'danger')
        except Exception as e:
            flash(f'Erro ao adicionar técnico: {str(e)}', 'danger')
        
        return render_template('administrador/form_tecnico.html', 
                               tecnico=request.form,
                               titulo="Adicionar Novo Técnico",
                               acao="Adicionar")
        
    return render_template('administrador/form_tecnico.html', 
                           titulo="Adicionar Novo Técnico", 
                           acao="Adicionar",
                           tecnico=None)

# NOVA ROTA: Editar técnico individual
@app.route('/admin/tecnicos/editar/<int:id_tecnico>', methods=['GET', 'POST'])
def editar_tecnico(id_tecnico):
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo_tecnico = request.form.get('tipo_tecnico')
        ativo = 1 if request.form.get('ativo') == '1' else 0 # Checkbox 'ativo'

        if not nome or not tipo_tecnico:
            flash('Nome e Tipo do Técnico são obrigatórios.', 'warning')
            # Para repopular o formulário em caso de erro, precisamos dos dados atuais do técnico
            cursor.execute("SELECT * FROM tecnicos WHERE id = ?", (id_tecnico,))
            tecnico_data = cursor.fetchone()
            if not tecnico_data: # Segurança, caso o técnico tenha sido deletado enquanto editava
                flash('Técnico não encontrado.', 'danger')
                return redirect(url_for('listar_tecnicos'))
            
            # Cria um dicionário com os dados do formulário para repopular
            form_data_com_erro = {
                'id': id_tecnico, # Mantém o ID original
                'nome': nome,
                'tipo_tecnico': tipo_tecnico,
                'ativo': ativo
            }
            return render_template('administrador/form_tecnico.html', 
                                   tecnico=form_data_com_erro, 
                                   titulo=f"Editar Técnico: {tecnico_data['nome']}", # Usa o nome original no título
                                   acao="Editar")

        try:
            # Verifica se o novo nome já existe para outro técnico
            cursor.execute("SELECT id FROM tecnicos WHERE nome = ? AND id != ?", (nome, id_tecnico))
            outro_tecnico_com_mesmo_nome = cursor.fetchone()
            if outro_tecnico_com_mesmo_nome:
                flash('Já existe outro técnico com este nome.', 'danger')
            else:
                cursor.execute("""
                    UPDATE tecnicos 
                    SET nome = ?, tipo_tecnico = ?, ativo = ?
                    WHERE id = ?
                """, (nome, tipo_tecnico, ativo, id_tecnico))
                conn.commit()
                flash('Técnico atualizado com sucesso!', 'success')
                return redirect(url_for('listar_tecnicos'))
        
        except sqlite3.Error as e:
            flash(f'Erro de banco de dados ao atualizar técnico: {str(e)}', 'danger')
        except Exception as e:
            flash(f'Erro ao atualizar técnico: {str(e)}', 'danger')

        # Se chegou aqui (erro de integridade ou outro erro de DB), repopula o form
        # com os dados que o usuário tentou submeter.
        form_data_com_erro = {
            'id': id_tecnico,
            'nome': nome,
            'tipo_tecnico': tipo_tecnico,
            'ativo': ativo
        }
        # Precisamos do nome original para o título em caso de erro de duplicação de nome
        cursor.execute("SELECT nome FROM tecnicos WHERE id = ?", (id_tecnico,))
        nome_original_tecnico = cursor.fetchone()['nome'] if cursor.fetchone() else "ID "+str(id_tecnico)

        return render_template('administrador/form_tecnico.html', 
                               tecnico=form_data_com_erro, 
                               titulo=f"Editar Técnico: {nome_original_tecnico}",
                               acao="Editar")

    # Método GET: Carregar dados do técnico para edição
    cursor.execute("SELECT * FROM tecnicos WHERE id = ?", (id_tecnico,))
    tecnico_data = cursor.fetchone()
    # conn.close() # Gerenciado pelo teardown_appcontext

    if not tecnico_data:
        flash('Técnico não encontrado.', 'danger')
        return redirect(url_for('listar_tecnicos'))

    return render_template('administrador/form_tecnico.html', 
                           tecnico=tecnico_data, 
                           titulo=f"Editar Técnico: {tecnico_data['nome']}",
                           acao="Editar")
# --- ADMIN - Gerenciamento de Locais ---
@app.route('/admin/locais')
def listar_locais():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM locais ORDER BY nome")
    locais = cursor.fetchall()
    conn.close()
    return render_template('administrador/listar_locais.html', locais=locais)

@app.route('/admin/locais/adicionar', methods=['GET', 'POST'])
def adicionar_local():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        ativo = 1 if request.form.get('ativo') == '1' else 0

        if not nome:
            flash('O nome do local é obrigatório.', 'danger')
            return render_template('administrador/form_local.html', local=None) # Retorna para o form com dados preenchidos

        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO locais (nome, descricao, ativo) VALUES (?, ?, ?)",
                (nome, descricao, ativo)
            )
            conn.commit()
            flash('Local adicionado com sucesso!', 'success')
            return redirect(url_for('listar_locais'))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash('Já existe um local com este nome.', 'danger')
        except Exception as e:
            conn.rollback()
            flash(f'Erro ao adicionar local: {str(e)}', 'danger')
        finally:
            if conn:
                conn.close()
        # Em caso de erro, recarrega o formulário com os dados inseridos
        return render_template('administrador/form_local.html', local={'nome': nome, 'descricao': descricao, 'ativo': ativo})


    return render_template('administrador/form_local.html', local=None) # Para GET request

@app.route('/admin/locais/editar/<int:id>', methods=['GET', 'POST'])
def editar_local(id):
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db() # Mova a obtenção da conexão para o início
    cursor = conn.cursor()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        ativo = 1 if request.form.get('ativo') == '1' else 0

        if not nome:
            flash('O nome do local é obrigatório.', 'danger')
            # Para manter os dados no formulário em caso de erro no POST
            cursor.execute("SELECT * FROM locais WHERE id = ?", (id,))
            local_data = cursor.fetchone()
            conn.close()
            if not local_data:
                flash('Local não encontrado.', 'danger')
                return redirect(url_for('listar_locais'))
            # Atualiza o dicionário com os dados do formulário para repopular
            local_dict_for_template = dict(local_data)
            local_dict_for_template['nome'] = nome
            local_dict_for_template['descricao'] = descricao
            local_dict_for_template['ativo'] = ativo
            return render_template('administrador/form_local.html', local=local_dict_for_template)


        try:
            # Verifica se o novo nome já existe para outro ID
            cursor.execute("SELECT id FROM locais WHERE nome = ? AND id != ?", (nome, id))
            outro_local_com_mesmo_nome = cursor.fetchone()
            if outro_local_com_mesmo_nome:
                flash('Já existe outro local com este nome.', 'danger')
            else:
                cursor.execute(
                    "UPDATE locais SET nome = ?, descricao = ?, ativo = ? WHERE id = ?",
                    (nome, descricao, ativo, id)
                )
                conn.commit()
                flash('Local atualizado com sucesso!', 'success')
                conn.close() # Fechar conexão aqui após o commit bem-sucedido
                return redirect(url_for('listar_locais'))
        except sqlite3.Error as e: # Usar sqlite3.Error para ser mais específico
            conn.rollback()
            flash(f'Erro ao atualizar local: {str(e)}', 'danger')
        # Não feche a conexão aqui se houve erro e você vai renderizar o template novamente

    # Para GET request ou se houve erro no POST e precisa recarregar com os dados do BD
    cursor.execute("SELECT * FROM locais WHERE id = ?", (id,))
    local_data = cursor.fetchone()
    conn.close() # Fechar a conexão após buscar os dados para o GET

    if not local_data:
        flash('Local não encontrado.', 'danger')
        return redirect(url_for('listar_locais'))
    
    return render_template('administrador/form_local.html', local=local_data)


@app.route('/admin/locais/remover/<int:id>')
def remover_local(id):
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    try:
        conn = get_db()
        cursor = conn.cursor()
        # Opcional: Verificar se o local está em uso antes de remover ou apenas desativar.
        # Por simplicidade, vamos remover.
        cursor.execute("DELETE FROM locais WHERE id = ?", (id,))
        conn.commit()
        flash('Local removido com sucesso!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao remover local: {str(e)}', 'danger')
    finally:
        if conn:
            conn.close()
    
    return redirect(url_for('listar_locais'))

# --- ADMIN - Gerenciamento de Usuários ---
@app.route('/admin/usuarios')
def listar_usuarios():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios ORDER BY nome")
    usuarios = cursor.fetchall()
    conn.close()

    return render_template('administrador/listar_usuarios.html', usuarios=usuarios)

@app.route('/admin/usuarios/novo', methods=['GET', 'POST'])
def novo_usuario():
    if session.get('tipo') not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    tipos_disponiveis_para_criacao = []
    if session.get('tipo') == 'master-admin':
        tipos_disponiveis_para_criacao = [
            {'value': 'solicitante', 'label': 'Solicitante'},
            {'value': 'manutencao', 'label': 'Login da Manutenção (Compartilhado)'}, # Rótulo ajustado
            {'value': 'admin', 'label': 'Administrador'}
        ]
    elif session.get('tipo') == 'admin':
        tipos_disponiveis_para_criacao = [
            {'value': 'solicitante', 'label': 'Solicitante'},
            {'value': 'manutencao', 'label': 'Login da Manutenção (Compartilhado)'} # Rótulo ajustado
        ]

    if request.method == 'POST':
        usuario_form = request.form.get('usuario', '').strip()
        senha_form = request.form.get('senha')
        tipo_novo_usuario = request.form.get('tipo')
        nome_form = request.form.get('nome', '').strip()
        email_form = request.form.get('email', '').strip()
        
        # A especialidade não é mais um campo para a tabela 'usuarios'
        # Ela será gerenciada na tabela 'tecnicos' separadamente.
        # O usuário do tipo 'manutencao' criado aqui é o login genérico.

        if not all([usuario_form, senha_form, tipo_novo_usuario, nome_form, email_form]):
            flash('Todos os campos marcados com * são obrigatórios.', 'warning')
            return render_template('administrador/novo_usuario.html', 
                                   tipos_disponiveis=tipos_disponiveis_para_criacao,
                                   submitted_data=request.form)
        
        tipos_valores_permitidos = [t['value'] for t in tipos_disponiveis_para_criacao]
        if tipo_novo_usuario not in tipos_valores_permitidos:
            flash('Você não tem permissão para criar este tipo de usuário.', 'danger')
            return render_template('administrador/novo_usuario.html', 
                                   tipos_disponiveis=tipos_disponiveis_para_criacao,
                                   submitted_data=request.form)

        # Se o tipo for 'manutencao', não precisamos mais pegar 'especialidade' aqui,
        # pois o usuário 'manutencao' é genérico. As especialidades estão na tabela 'tecnicos'.

        try:
            conn = get_db()
            cursor = conn.cursor()
            # Query INSERT SEM a coluna 'especialidade'
            cursor.execute('''
                INSERT INTO usuarios (usuario, senha, tipo, nome, email, ativo)
                VALUES (?, ?, ?, ?, ?, 1)
            ''', (usuario_form, generate_password_hash(senha_form), tipo_novo_usuario, nome_form, email_form))
            conn.commit()
            flash('Usuário criado com sucesso!', 'success')
            return redirect(url_for('listar_usuarios'))
        except sqlite3.IntegrityError:
            flash('Nome de usuário ou email já cadastrado.', 'danger')
        except Exception as e:
            flash(f'Erro ao criar usuário: {str(e)}', 'danger')
        
        return render_template('administrador/novo_usuario.html', 
                               tipos_disponiveis=tipos_disponiveis_para_criacao,
                               submitted_data=request.form)

    return render_template('administrador/novo_usuario.html', 
                           tipos_disponiveis=tipos_disponiveis_para_criacao,
                           submitted_data=None) # Para GET, não há dados submetidos

@app.route('/admin/usuarios/editar/<int:id_usuario_alvo>', methods=['GET', 'POST'])
def editar_usuario(id_usuario_alvo):
    editor_tipo = session.get('tipo')
    editor_id = session.get('user_id')

    if editor_tipo not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios WHERE id = ?", (id_usuario_alvo,))
    usuario_alvo = cursor.fetchone()

    if not usuario_alvo:
        flash('Usuário não encontrado!', 'danger')
        return redirect(url_for('listar_usuarios'))

    # --- Lógica de Permissão para Acessar a Página de Edição (GET) e para Salvar (POST) ---
    pode_editar = False
    tipos_disponiveis_para_atribuicao = []

    if editor_tipo == 'master-admin':
        pode_editar = True
        # Master-admin pode atribuir estes tipos:
        tipos_disponiveis_para_atribuicao = [
            {'value': 'solicitante', 'label': 'Solicitante'},
            {'value': 'manutencao', 'label': 'Técnico de Manutenção'},
            {'value': 'admin', 'label': 'Administrador'}
        ]
        # Master-admin não pode rebaixar outro master-admin ou a si mesmo via este form simples
        if usuario_alvo['tipo'] == 'master-admin':
            tipos_disponiveis_para_atribuicao = [{'value': 'master-admin', 'label': 'Master Administrador'}] # Só pode ser master-admin
            if usuario_alvo['id'] == editor_id and sum(1 for row in cursor.execute("SELECT 1 FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1")) <= 1:
                 # Impede o único master-admin ativo de mudar seu próprio tipo se for o único
                 tipos_disponiveis_para_atribuicao = [{'value': 'master-admin', 'label': 'Master Administrador'}]


    elif editor_tipo == 'admin':
        if usuario_alvo['tipo'] == 'master-admin':
            pode_editar = False # Admin não pode editar master-admin
            flash('Administradores não podem editar usuários Master Administrador.', 'warning')
        elif usuario_alvo['tipo'] == 'admin' and usuario_alvo['id'] != editor_id:
            pode_editar = False # Admin não pode editar outros admins
            flash('Administradores não podem editar outros Administradores.', 'warning')
        elif usuario_alvo['id'] == editor_id: # Admin editando a si mesmo
            pode_editar = True # Pode editar seus próprios dados, exceto o tipo
            tipos_disponiveis_para_atribuicao = [{'value': 'admin', 'label': 'Administrador'}] # Tipo fixo
        else: # Admin editando solicitante ou manutencao
            pode_editar = True
            tipos_disponiveis_para_atribuicao = [
                {'value': 'solicitante', 'label': 'Solicitante'},
                {'value': 'manutencao', 'label': 'Técnico de Manutenção'}
            ]
            # Se o usuário alvo já é admin (e o editor não é o mesmo admin), essa condição não deveria ser alcançada
            # por causa do 'elif' anterior. Mas para garantir:
            if usuario_alvo['tipo'] == 'admin':
                 tipos_disponiveis_para_atribuicao = [{'value': 'admin', 'label': 'Administrador'}]


    if not pode_editar and request.method == 'GET': # Redireciona no GET se não puder editar
         return redirect(url_for('listar_usuarios'))


    if request.method == 'POST':
        if not pode_editar: # Dupla verificação para o POST
            flash('Você não tem permissão para modificar este usuário.', 'danger')
            return redirect(url_for('listar_usuarios'))

        nome_form = request.form.get('nome','').strip()
        usuario_form = request.form.get('usuario','').strip() # Usuário (login)
        email_form = request.form.get('email','').strip()
        tipo_form = request.form.get('tipo')
        ativo_form = 1 if request.form.get('ativo') == 'on' else 0 # Checkbox 'on' ou ausente
        especialidade_form = request.form.get('especialidade') if tipo_form == 'manutencao' else None

        # Validação básica
        if not all([nome_form, usuario_form, email_form, tipo_form]):
            flash('Campos Nome, Usuário, Email e Tipo são obrigatórios.', 'warning')
            # Recarrega com os dados do formulário, não do banco, para o usuário corrigir
            usuario_alvo_temp_form = dict(usuario_alvo)
            usuario_alvo_temp_form.update(request.form.to_dict())
            return render_template('administrador/editar_usuario.html', 
                                   usuario=usuario_alvo_temp_form, 
                                   tipos_disponiveis=tipos_disponiveis_para_atribuicao,
                                   pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))


        # Validação do tipo selecionado
        tipos_valores_permitidos = [t['value'] for t in tipos_disponiveis_para_atribuicao]
        if tipo_form not in tipos_valores_permitidos:
            flash('Você não tem permissão para atribuir este tipo de usuário.', 'danger')
            # Recarrega como acima
            usuario_alvo_temp_form = dict(usuario_alvo)
            usuario_alvo_temp_form.update(request.form.to_dict())
            return render_template('administrador/editar_usuario.html', 
                                   usuario=usuario_alvo_temp_form, 
                                   tipos_disponiveis=tipos_disponiveis_para_atribuicao,
                                   pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))

        if tipo_form == 'manutencao' and not especialidade_form:
            flash('Especialidade é obrigatória para Técnico de Manutenção.', 'warning')
            # Recarrega como acima
            usuario_alvo_temp_form = dict(usuario_alvo)
            usuario_alvo_temp_form.update(request.form.to_dict())
            return render_template('administrador/editar_usuario.html', 
                                   usuario=usuario_alvo_temp_form, 
                                   tipos_disponiveis=tipos_disponiveis_para_atribuicao,
                                   pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))
        
        # Lógica para impedir que o único master-admin ativo se desative ou mude de tipo crítico
        if usuario_alvo['tipo'] == 'master-admin' and usuario_alvo['id'] == editor_id:
            if not ativo_form:
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1 AND id != ?", (editor_id,))
                outros_master_admins_ativos = cursor.fetchone()[0]
                if outros_master_admins_ativos == 0:
                    flash('Não é possível desativar o único Master Administrador ativo.', 'danger')
                    ativo_form = 1 # Força de volta para ativo
            if tipo_form != 'master-admin':
                 flash('Master Administradores não podem mudar seu próprio tipo desta forma.', 'danger')
                 tipo_form = 'master-admin' # Força de volta para master-admin


        try:
            cursor.execute('''
                UPDATE usuarios 
                SET nome = ?, usuario = ?, email = ?, tipo = ?, especialidade = ?, ativo = ?
                WHERE id = ?
            ''', (nome_form, usuario_form, email_form, tipo_form, especialidade_form, ativo_form, id_usuario_alvo))
            conn.commit()
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('listar_usuarios'))
        except sqlite3.IntegrityError:
            # conn.rollback()
            flash('Nome de usuário ou email já existe para outro usuário.', 'danger')
        except Exception as e:
            # conn.rollback()
            flash(f'Erro ao atualizar usuário: {str(e)}', 'danger')
        
        # Se chegou aqui, houve erro no update, repopula o form
        usuario_alvo_temp_form = dict(usuario_alvo)
        usuario_alvo_temp_form.update(request.form.to_dict()) # Pega os valores atuais do form
        return render_template('administrador/editar_usuario.html', 
                               usuario=usuario_alvo_temp_form, 
                               tipos_disponiveis=tipos_disponiveis_para_atribuicao,
                               pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))


    # Método GET (se pode_editar for True)
    return render_template('administrador/editar_usuario.html', 
                           usuario=usuario_alvo, 
                           tipos_disponiveis=tipos_disponiveis_para_atribuicao,
                           pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))

@app.route('/admin/usuarios/alterar_senha/<int:id_usuario_alvo>', methods=['POST'])
def alterar_senha_usuario(id_usuario_alvo):
    editor_tipo = session.get('tipo')
    editor_id = session.get('user_id')

    if editor_tipo not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))
    nova_senha = request.form.get('nova_senha')
    if not nova_senha:
        flash('Nova senha não pode ser vazia.', 'warning')
        return redirect(url_for('editar_usuario', id_usuario_alvo=id_usuario_alvo)) 
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE usuarios 
            SET senha = ?
            WHERE id = ?
        ''', (generate_password_hash(nova_senha), id_usuario_alvo))
        conn.commit()
        flash('Senha alterada com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao alterar senha: {str(e)}', 'danger')
    
    return redirect(url_for('editar_usuario', id_usuario_alvo=id_usuario_alvo))

@app.route('/admin/usuarios/remover/<int:id_usuario_alvo>')
def remover_usuario(id_usuario_alvo):
    editor_tipo = session.get('tipo')
    editor_id = session.get('user_id')

    if editor_tipo not in ['admin', 'master-admin']:
        flash('Acesso não autorizado.', 'danger')
        return redirect(url_for('login'))

    if id_usuario_alvo == editor_id:
        flash('Você não pode remover a si mesmo!', 'danger')
        return redirect(url_for('listar_usuarios'))

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM usuarios WHERE id = ?", (id_usuario_alvo,))
    usuario_alvo = cursor.fetchone()

    if not usuario_alvo:
        flash('Usuário a ser removido não encontrado.', 'warning')
        return redirect(url_for('listar_usuarios'))

    pode_remover = False
    mensagem_erro_permissao = 'Você não tem permissão para remover este tipo de usuário.'

    if editor_tipo == 'master-admin':
        # Master-admin não pode remover outro master-admin se for o único ativo restante
        if usuario_alvo['tipo'] == 'master-admin':
            if usuario_alvo['id'] == editor_id: # Já coberto acima, mas por segurança
                pode_remover = False
                mensagem_erro_permissao = 'Você não pode remover a si mesmo!'
            else:
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1 AND id != ?", (id_usuario_alvo,))
                count_outros_master_admins_ativos = cursor.fetchone()[0]
                if count_outros_master_admins_ativos >= 1:
                    pode_remover = True
                else:
                    cursor.execute("SELECT COUNT(*) FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1")
                    total_master_admins_ativos = cursor.fetchone()[0]
                    if total_master_admins_ativos > 1:
                        pode_remover = True
                    else:
                        pode_remover = False
                        mensagem_erro_permissao = 'Não é possível remover o único Master Administrador ativo ou deixar o sistema sem Master Administradores ativos.'
        else:
            # Master-admin pode remover admin, manutencao, solicitante
            pode_remover = True
    
    elif editor_tipo == 'admin':
        # Admin NÃO PODE remover master-admin
        if usuario_alvo['tipo'] == 'master-admin':
            pode_remover = False
            mensagem_erro_permissao = 'Administradores não podem remover usuários Master Administrador.'
        # Admin NÃO PODE remover outros admins
        elif usuario_alvo['tipo'] == 'admin':
            pode_remover = False
            mensagem_erro_permissao = 'Administradores não podem remover outros Administradores.'
        # Admin PODE remover solicitante e manutencao
        elif usuario_alvo['tipo'] in ['solicitante', 'manutencao']:
            pode_remover = True
        else:
            pode_remover = False # Por segurança, nega outros casos

    if not pode_remover:
        flash(mensagem_erro_permissao, 'danger')
        return redirect(url_for('listar_usuarios'))

    try:
        cursor.execute("SELECT COUNT(*) FROM ordens_servico WHERE solicitante_id = ? OR tecnico_id = ?", (id_usuario_alvo, id_usuario_alvo))
        if cursor.fetchone()[0] > 0:
            flash(f"Não é possível remover o usuário (ID: {id_usuario_alvo}) pois ele está associado a Ordens de Serviço. Considere desativá-lo.", 'warning')
            return redirect(url_for('listar_usuarios'))

        cursor.execute("SELECT COUNT(*) FROM registros_manutencao_direta WHERE criado_por_id = ? OR concluido_por_admin_id = ?", (id_usuario_alvo, id_usuario_alvo))
        if cursor.fetchone()[0] > 0:
            flash(f"Não é possível remover o usuário (ID: {id_usuario_alvo}) pois ele está associado a Registros de Manutenção. Considere desativá-lo.", 'warning')
            return redirect(url_for('listar_usuarios'))

        cursor.execute("DELETE FROM participantes_os WHERE tecnico_id = ?", (id_usuario_alvo,))
        cursor.execute("DELETE FROM participantes_registro_direto WHERE tecnico_id = ?", (id_usuario_alvo,))
 
        
        cursor.execute("DELETE FROM usuarios WHERE id = ?", (id_usuario_alvo,))
        conn.commit()
        flash('Usuário removido com sucesso!', 'success')
    except sqlite3.IntegrityError as e:
        flash(f"Erro de integridade ao remover usuário: {str(e)}. Verifique se o usuário está referenciado em outras tabelas.", 'danger')
    except Exception as e:
        flash(f'Erro ao remover usuário: {str(e)}', 'danger')
    
    return redirect(url_for('listar_usuarios'))

# No final do arquivo, antes do app.run()
app.teardown_appcontext(close_db)

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)